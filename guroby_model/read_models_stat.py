import openpyxl

book = openpyxl.Workbook()
sheet = book.active

x_shift = 2
y_shift = 3

#  запись 
I = range(10)
K = range(1, 4)
for i in I:

    sheet[y_shift + i][0].value = f"exp {i + 1}"


    max_obj_val = 0.0
    max_k = None
    for k in K:
        # print(f"consol_sol_{i + 1}_time_3600_K_{k + 1}_EX_{i + 1}_ver4.1.txt")
        file_name = f"consol_sol_{i + 1}_time_3600_K_{k + 1}_EX_{i + 1}_ver4.1.txt"

        try:
            with open(file_name, 'r' ) as file:
                read = file.readlines()
                
        except:
        
            print(f"File don't esixt K_{k + 1}_EX_{i + 1}")
            continue
        


        obj_val = float(read[-4][read[-4].find(':') + len(':')+ 1:].split(' ')[0])
        up_bound = float(read[-1][read[-1].find('best bound') + len('best bound') + 1:].split(' ')[0][:-1])
        gap = read[-1][read[-1].find('gap') + len('gap') + 1:].split(' ')[0][:-2]
        

        if obj_val > max_obj_val:
            max_obj_val = obj_val
            max_k = k    

        sheet.cell(row = 2, column =x_shift + k*4  ).value = "obj_val"
        sheet.cell(row = 2, column =x_shift + k*4 + 1 ).value = "up_bound"
        sheet.cell(row = 2, column =x_shift + k*4 + 2 ).value = "gap %"

        sheet.cell(row = y_shift + i, column =x_shift + k*4  ).value = obj_val
        sheet.cell(row = y_shift + i, column =x_shift + k*4 + 1 ).value = up_bound
        sheet.cell(row = y_shift + i, column =x_shift + k*4 + 2 ).value = gap

        # sheet[y_shift + i][1 + k*4 ].value = str(up_bound)
        # sheet[y_shift + i][1 + k*4 ].value = str(gap)

    k = max_k

    # print(f"consol_sol_{i + 1}_time_3600_K_{k + 1}_EX_{i + 1}_ver4.1.txt")
    file_name = f"consol_sol_{i + 1}_time_3600_K_{k + 1}_EX_{i + 1}_ver4.1.txt"

    try:
        with open(file_name, 'r' ) as file:
            read = file.readlines()
            
    except:
    
        print(f"File don't esixt K_{k + 1}_EX_{i + 1}")
        continue
    


    obj_val = float(read[-4][read[-4].find(':') + len(':')+ 1:].split(' ')[0])
    up_bound = float(read[-1][read[-1].find('best bound') + len('best bound') + 1:].split(' ')[0][:-1])
    gap = read[-1][read[-1].find('gap') + len('gap') + 1:].split(' ')[0][:-2]
    
    sheet.cell(row = 2, column =x_shift + 10*4  ).value = "obj_val"
    sheet.cell(row = 2, column =x_shift + 10*4 + 1 ).value = "up_bound"
    sheet.cell(row = 2, column =x_shift + 10*4 + 2 ).value = "gap %"

    sheet.cell(row = y_shift + i, column =x_shift + 10*4  ).value = obj_val
    sheet.cell(row = y_shift + i, column =x_shift + 10*4 + 1 ).value = up_bound
    sheet.cell(row = y_shift + i, column =x_shift + 10*4 + 2 ).value = gap







book.save("Gurobi_exp.xlsx")
book.close()